from openpyxl import load_workbook
from openpyxl import Workbook

#Opened and Initialization
student_names = load_workbook("csattendance.xlsx")
students_attendance = student_names.active
workbook = Workbook()
updated_attendance = workbook.active
updated_attendance.title = "CS Club Attendance"
voting_eligibility = 2
running_eligibility = 2

#Titles of Rows
updated_attendance["A1"] = "Names"
updated_attendance["B1"] = "Total Meetings Attended"
updated_attendance["C1"] = "Can Vote?"
updated_attendance["D1"] = "Can Run?"

#Store student_names column as a list
names_list = [name.value for name in students_attendance["E"]]
del names_list[0]
names_list_unique = set(names_list)
number_of_names = len(names_list_unique)


#Adds names and count of each into final attendance sheet
for rows in range(2, number_of_names+2):
    if names_list[rows-2] not in updated_attendance["A"]:
        updated_attendance["A" + str(rows)] = names_list[rows-2]
        updated_attendance["B" + str(rows)] = names_list.count(names_list[rows-2])
        if (names_list.count(names_list[rows-2]) >= voting_eligibility):
            updated_attendance["C"+str(rows)] = True
        else:
            updated_attendance["C"+str(rows)] = False
        if (names_list.count(names_list[rows-2]) >= running_eligibility):
            updated_attendance["D"+str(rows)] = True
        else:
            updated_attendance["D"+str(rows)] = False

            
    

    








#Save the workbook
workbook.template = True
workbook.save("clubattendancefinal.xltx")



